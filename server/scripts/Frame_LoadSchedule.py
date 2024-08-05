import os,sys
import openpyxl
import teradatasql
import logging
import datetime
import getConfig as gc
import pandas as pd
import Frame_Util as fg
xx=sys.argv[0]

# Validamos que ingresen parametros Layout y Fecha
if len(sys.argv) <2:
   print("Parametro no Ingresado...!")
   print("Uso: python {0} fileSchedule".format(sys.argv[0]))
   sys.exit()

if fg.procesosConcurrentes("Frame_LoadSchedule.py") > 1:
   logging.info('Otro usuario esta ejecutando la carga de schedule espere un memento y vuelva a intentar...!')
   print('Ya se esta ejecutando en otra Terminal "Frame_LoadSchedule.py"...')
   sys.exit()

xParam1 = sys.argv[1]
fec1 = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
fecha=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
FileLog = os.path.join(gc.path_l, "RS_"+ xParam1 + fecha + ".log")
print("FileLog: ",FileLog)
logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s : %(levelname)s : %(message)s',
                    filename = FileLog,
                    filemode = 'w',)

os.chdir(gc.path_e)


logging.debug('Inicio de programa')
logging.info('Inicio conexion a Teradata...')


#conectarse a Teradata
try:
    cntd = teradatasql.connect(None, host=gc.hostd, user=gc.usrtd, password=gc.pwdtd)
    curtd = cntd.cursor()

except:
    logging.warning('No se ha podido realizar la conexion.')
    logging.warning(sys.exc_info()[0:])
    sys.exit()


# -------    Programa Principal    ------ #
def main_py():
    try:
        sql_stm= "SELECT coalesce(MAX(ScheduleCD),0) FROM {}.TB_SCHEDULE".format(gc.bdcfg)
        ScheduleCD = read_sql2(sql_stm)
        xFileExcel = str(xParam1)
        logging.info('Leyendo archivo excel:' + xFileExcel)
        logging.info('Extrayendo información para carga...')
        reg_Schedule(xFileExcel,ScheduleCD)
    except:
        logging.warning(' No se ha podido realizar la conexion.')
        curtd.execute(sql_stm)
        logging.warning(sys.exc_info()[0:])
        sys.exit()

def reg_Schedule(fileExcel,sql_campo):
    try:
        sw = True
        wb1 = openpyxl.load_workbook(fileExcel, data_only=True)
        wb1.active = 0
        h1 = wb1.active
#        UnidadFrecCD = ''
        datos=[]
        for i in range(3,540):
            if h1.cell(i,2).value is None :
                break
            else:
                if h1.cell(i,3).value in "FTP, FLoad, NzLoad, RNTdt, RNNz, FxFlFL":
                    sql_stm = "SELECT LayoutCD,PeriodicidadCD FROM {}.VW_LAYOUT \
                    where NombreLayout= '{}'".format(gc.bdcfg, h1.cell(i,2).value)

                elif h1.cell(i,3).value in "SPTdt, SPNz":
                    sql_stm = "SELECT StoreCD, UnidadFrecCd FROM {}.TB_STOREPROCEDURE where CodigoSP= '{}' \
                    AND estadoSP=1".format(gc.bdcfg, h1.cell(i,2).value)

                elif h1.cell(i,3).value in "ExTdt, ExNz, ExOra":
                    sql_stm = "SELECT LayoutExtrID, UnidadFrecCD FROM {}.VW_LAYOUT_EXTR \
                    where NombLayout='{}' ".format(gc.bdcfg, h1.cell(i,2).value)

                elif h1.cell(i,3).value in "ExIfx":
                    sql_stm = "SELECT Tabla_ID,101 as UnidadFrecCD FROM {}.TB_GEN_TABLAS where Archivo_Nombre= '{}' AND \
                    Tipo_archivo_id=1 and fuente= 'CMS' AND PRODUCCION_IND='SI'".format(gc.bdcfg, h1.cell(i,2).value)

                elif h1.cell(i,3).value in "ExGst":
                    sql_stm = "SELECT Tabla_ID,101 as UnidadFrecCD FROM {}.TB_GEN_TABLAS where Archivo_Nombre= '{}' AND \
                    Tipo_archivo_id=1 and fuente= 'GESTEL' AND PRODUCCION_IND='SI'".format(gc.bdcfg, h1.cell(i,2).value)

                elif h1.cell(i,3).value in "JnFlFL":
                    sql_stm = "SELECT RoutineCd, UnidadFrecCD FROM {}.TB_ROUTINE_PARAMETROS_SO WHERE RoutineCodigo = '{}' AND \
                    RoutineTipoCd = '1'".format(gc.bdcfg, h1.cell(i,2).value)
                
                elif h1.cell(i,3).value in "GRNz":
                    sql_stm = "SELECT LayoutCD, UnidadFrecCD FROM {}.VW_GEN_REPORTE_NORMA \
                    WHERE CodReporte = '{}'".format(gc.bdcfg, h1.cell(i,2).value)


                logging.info("L0 i:{} TipoLayout: {}, SQL:{}".format(i,h1.cell(i,3).value,sql_stm))
                sql_campo = sql_campo +1
                row = read_sql1(sql_stm)
                LayoutCD = str(row[0])
                UnidadFrecCD = str(row[1])
                sql_stm = "SELECT ScheduleCD FROM {}.TB_SCHEDULE WHERE LayoutCD= {} \
                AND TipSchdCD='{}'" .format(gc.bdcfg,str(LayoutCD), h1.cell(i,12).value)
                # Si existe Realizamos Update
                existe= read_sql2(sql_stm)
#                if (existe<1):

                if h1.cell(i,9).value is None :
                    ScheduleCDPred = 0
                else:
                    if h1.cell(i,10).value in "FTP, FLoad, NzLoad, RNTdt, RNNz, FxFlFL":
                        sql_stm = "SELECT LayoutCD FROM {}.VW_LAYOUT where NombreLayout= '{}'".format(gc.bdcfg, h1.cell(i,9).value)
                    elif h1.cell(i,10).value in "SPTdt, SPNz, DUMMY":
                        sql_stm = "SELECT StoreCD FROM {}.TB_STOREPROCEDURE where CodigoSP= '{}' \
                        AND estadoSP=1".format(gc.bdcfg, h1.cell(i,9).value)

                    elif h1.cell(i,10).value in "ExTdt, ExNz, ExOra":
                        sql_stm = "SELECT LayoutExtrID FROM {}.VW_LAYOUT_EXTR where NombLayout= '{}'".format(gc.bdcfg, h1.cell(i,9).value)

                    elif h1.cell(i,10).value in "ExIfx":
                        sql_stm = "SELECT Tabla_ID FROM {}.TB_GEN_TABLAS where Archivo_Nombre= '{}' AND \
                        Tipo_archivo_id=1 and fuente= 'CMS' AND PRODUCCION_IND='SI'".format(gc.bdcfg, h1.cell(i,9).value)

                    elif h1.cell(i,10).value in "ExGst":
                        sql_stm = "SELECT Tabla_ID FROM {}.TB_GEN_TABLAS where Archivo_Nombre= '{}' AND \
                        Tipo_archivo_id=1 and fuente= 'GESTEL' AND PRODUCCION_IND='SI'".format(gc.bdcfg, h1.cell(i,9).value)

                    elif h1.cell(i,10).value in "JnFlFL":
                        sql_stm = "SELECT RoutineCd FROM {}.TB_ROUTINE_PARAMETROS_SO WHERE RoutineCodigo = '{}' AND \
                        RoutineTipoCD = '1'".format(gc.bdcfg, h1.cell(i,9).value)
                    
                    elif h1.cell(i,10).value in "GRNz":
                        sql_stm = "SELECT LayoutCD FROM {}.VW_GEN_REPORTE_NORMA WHERE CodReporte = '{}'".format(gc.bdcfg, h1.cell(i,9).value)

                    logging.info("TipoLayout: {}, SQL:{}".format(h1.cell(i,3).value,sql_stm))
                    LayoutCDP = read_sql2(sql_stm)

                    sql_stm = "SELECT ScheduleCD FROM {}.TB_SCHEDULE WHERE LayoutCD='{}'AND TipSchdCD='{}'" .format(gc.bdcfg, str(LayoutCDP),+  h1.cell(i,13).value)
                    ScheduleCDPred = read_sql2(sql_stm)

                xListaValores = str(h1.cell(i,7).value)                    
                xValorFrecuencia = int(str(h1.cell(i,7).value).split(',')[0])

                if (existe<1):
                    #Insertamos Schedule
                    datos=[LayoutCD, UnidadFrecCD, h1.cell(i,4).value, h1.cell(i,5).value, xValorFrecuencia, h1.cell(i,11).value, \
                           h1.cell(i,6).value, fec1, h1.cell(i,8).value, ScheduleCDPred, h1.cell(i,12).value,sql_campo, h1.cell(i,14).value, xListaValores]
                    stm_sql="INSERT INTO {}.TB_SCHEDULE\
                    (LayoutCD, UnidadFrecCD, FecIni, HorIni, ValorFrecuencia, FlagIndFin,\
                    HorFin, FecCreaTS, HorMaxDura, ScheduleCDPred, TipSchdCD, ScheduleCD, Predecesores, ValorFrecuenciaAdic)\
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)".format(gc.bdcfg)
                    curtd.execute(stm_sql,datos)
                    logging.info('Se registro Schedule -> {} {}'.format( h1.cell(i,2).value, h1.cell(i,3).value))
                else:
                    #Generamos Historia
                    stm_sql = """INSERT INTO {0}.TB_SCHEDULE_HIST
                    SELECT t.*, Current_Timestamp(0) FROM {0}.TB_SCHEDULE t WHERE ScheduleCD = {1}""".format(gc.bdcfg,existe)
                    curtd.execute(stm_sql)
                    #Actualizamos Schedule
                    datos=[ UnidadFrecCD, h1.cell(i,4).value, h1.cell(i,5).value, xValorFrecuencia, h1.cell(i,11).value, \
                           h1.cell(i,6).value, fec1, h1.cell(i,8).value, ScheduleCDPred, h1.cell(i,14).value, xListaValores]
                    stm_sql = "UPDATE  {}.TB_SCHEDULE\
                    SET UnidadFrecCD=?, FecIni=?, HorIni=?, ValorFrecuencia=?, FlagIndFin=?, HorFin=?, FecCreaTS=?\
                    ,HorMaxDura=?, ScheduleCDPred=?, Predecesores=?, ValorFrecuenciaAdic=? \
                    where ScheduleCD = {} ".format(gc.bdcfg,existe)
                    logging.info("Upt stm_sql: {}".format(stm_sql))
                    curtd.execute(stm_sql,datos)
                    logging.info('Schedule ya registrado-> {} {}'.format( h1.cell(i,2).value, h1.cell(i,3).value))

        wb1.close()
        logging.info('Fin de registro schedule')
    except:
        logging.warning('Error en reg_Schedule.')
        logging.error("Error inesperado: %s", sys.exc_info()[0:])
        sys.exit()


def read_sql2(query):
    try:
        curtd.execute(query)
        row = curtd.fetchone()
        if row:
           return(row[0])
        else:
           return 0
    except:
        logging.warning('Error en read_sql2.\n',query)
        logging.error("Error inesperado: %s", sys.exc_info()[0:])
        sys.exit()


def read_sql1(query):
    try:
        curtd.execute(query)
        row = curtd.fetchone()
        return(row)
    except:
        logging.warning('Error en read_sql2.\n',query)
        logging.error("Error inesperado: %s", sys.exc_info()[0:])
        sys.exit()



if __name__ == "__main__":
    main_py();

