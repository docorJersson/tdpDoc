import os,sys
import openpyxl
import teradatasql
import logging
import datetime
import getConfig as gc
import pandas as pd
xx=sys.argv[0]

if len(sys.argv) <2:
   print("Parametro no Ingresado...!")
   print("Uso: python {0} fileExcel.xlsx".format(sys.argv[0]))
   sys.exit()

xParam1 = sys.argv[1]
fec1 = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
fecha=datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
FileLog = os.path.join(gc.path_l, "RSP_"+ xParam1 +"_"+ fecha + ".log")
logging.basicConfig(level=logging.DEBUG,
                    format='%(asctime)s : %(levelname)s : %(message)s',
                    filename = FileLog,
                    filemode = 'w',)

os.chdir(gc.path_e)

host, username, password = gc.hostd, gc.usrtd, gc.pwdtd

logging.debug('Inicio de programa')
logging.info('Inicio conexion a Teradata...')
#conectarse a Teradata
try:
    cntd = teradatasql.connect(None, host=gc.hostd, user=gc.usrtd, password=gc.pwdtd)
    curtd = cntd.cursor()
    logging.info("Conectado a Teradata")

except:
    logging.warning('No se ha podido realizar la conexion.')
    logging.warning(sys.exc_info()[0:])
    sys.exit()


# -------    Programa Principal    ------ #
def main_py():
    try:
       xFileExcel = str(xParam1)
       logging.info('1: Leyendo archivo excel:' + xFileExcel)
       logging.info('1: Extrayendo información para carga...')
       campo = read_excel2(xFileExcel)
       logging.info('1: Realizando Registro TB_STOREPROCEDURE ...')       
       ejecutaCarga(campo)
       

    except:
        logging.warning('1: No se ha podido realizar la conexion.')
        logging.warning(sys.exc_info()[0:])
        sys.exit()

def read_sql(query):
    try:
        curtd.execute(query)
        row = curtd.fetchone()
        return(row[0])
    except:
        logging.warning('2: Error en read_sql.\n',query)
        logging.error("2: Error inesperado: %s", sys.exc_info()[0:])
        sys.exit()


def read_excel2(fileExcel):
    try:
        sw = True
        wb1 = openpyxl.load_workbook(fileExcel, data_only=True)
        wb1.active = 0
        h1 = wb1.active
        datos=[]
        for i in range(3,540):
            if h1.cell(i,2).value is None :
                break
            else:
                if (sw):
                    sql_stm  = "SELECT coalesce(MAX(StoreCd)+1,1) FROM {}.TB_STOREPROCEDURE".format(gc.bdcfg)
                    StoreCd = read_sql(sql_stm)
                    sw = False
                else:
                    StoreCd = StoreCd +1

                sql_stm= "SELECT count(*)+1 FROM {}.TB_STOREPROCEDURE \
                where CodigoSP='{}' ".format(gc.bdcfg,h1.cell(i,2).value)
                version = read_sql(sql_stm)

                if version > 1:
                    logging.info('3: Store Procedure ya registrado  ')
                    logging.info('3: ...insertando nueva versión ')
                    sql_stm= "UPDATE {}.TB_STOREPROCEDURE \
                    SET estadoSP=0 \
                    WHERE CodigoSP='{}' ".format(gc.bdcfg,h1.cell(i,2).value)
                    curtd.execute(sql_stm)
                    
                    sql_stm= "SELECT StoreCD FROM {}.TB_STOREPROCEDURE where CodigoSP='{}'".format(gc.bdcfg,h1.cell(i,2).value)
                    StoreCd = read_sql(sql_stm)

                estado= 1

                datos.append([StoreCd, h1.cell(i,2).value, h1.cell(i,3).value, h1.cell(i,4).value, \
                              h1.cell(i,5).value,h1.cell(i,6).value, h1.cell(i,7).value,h1.cell(i,8).value,\
                              h1.cell(i,9).value, h1.cell(i,10).value,estado, version,fec1,h1.cell(i,12).value])
        wb1.close()
        return datos

    except:
      logging.warning('3: Error en read_excel2.')
      logging.error("3: Error inesperado: %s", sys.exc_info()[0:])
      sys.exit()


def ejecutaCarga(campo):

    sql_stm="insert into {}.TB_STOREPROCEDURE (StoreCD,CodigoSP,BaseDatosSP,UsuarioBD,EsquemaBD,NombreSP,DescripcionSP,Parametro1 \
    ,Parametro2,Parametro3,estadoSP,version,FECHACREACION, UnidadFrecCD) \
    values(?,?,?,?,?,?,?,?,?,?,?,?,?,?)".format(gc.bdcfg)
    logging.info('4: Inicio de registro en TB_STOREPROCEDURE: {} '.format(campo))
    curtd.executemany(sql_stm,campo)
    logging.info('4: ... Fin registro TB_STOREPROCEDURE ')



if __name__ == "__main__":
    main_py();

